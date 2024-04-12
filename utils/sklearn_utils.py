#!/usr/bin/env python
# -*-coding:utf-8 -*-
"""
@Project :   Study
@File    :   sklearn_utils.py
@Time    :   2024-04-10 13:54
@Author  :   Wu Xiaomin <>
@Version :   1.0
@License :   (C)Copyright 2024, Wu Xiaomin
@Desc    :   sklearn 相关工具类
"""
import os
from collections import defaultdict
from pathlib import Path

import pandas as pd
from joblib import load
from openpyxl import load_workbook
from sklearn.metrics import classification_report
from sklearn.model_selection import cross_val_score


def get_var_name(var):
    for name, value in locals().items():
        if value is var:
            return name
    else:
        raise ValueError(f"Variable {var} not found in local scope.")


class ClassifierZoo(object):

    def __init__(self, **kwargs):
        self.classifiers = []
        self.config = {
            "dump_path": "./",
            # 分类问题: accuracy、f1、f1_micro、f1_macro（这两个用于多分类的f1_score）、precision、recall、roc_auc
            "scoring": "f1_micro",
            "cv": 5
        }
        self.config.update(kwargs)
        self.train_report = defaultdict(list)

    def add_classifier(self, classifier):
        self.classifiers.append(classifier)

    @classmethod
    def from_model_path(cls, model_path: os.PathLike, **kwargs):
        instance = cls(**kwargs)
        for model in Path(model_path).glob("*.pkl"):
            instance.add_classifier(load(model))
        return instance

    def train(self, X, y):
        for cls in self.classifiers:
            cls_name = get_var_name(cls)
            cls.fit(X, y)
            scores = cross_val_score(cls, X, y, scoring=self.config["scoring"], cv=self.config["cv"])
            print(f"train {cls_name}, score_metric: {self.config['scoring']}")
            print(scores.mean())
            print(scores.std())
            self.train_report["classifier"].append(cls_name)
            self.train_report[self.config["scoring"] + "_mean"].append(scores.mean())
            self.train_report[self.config["scoring"] + "_std"].append(scores.std())
        print(f"train {len(self.classifiers)} classifiers complete.")
        pd.DataFrame(self.train_report).to_excel(self.config["dump_path"] + "/train_report.xlsx")

    def test(self, X_test, y_test):
        for cls in self.classifiers:
            cls_name = get_var_name(cls)
            y_pred = cls.predict(X_test)
            d = classification_report(y_test, y_pred, output_dict=True)
            summay_excel_report(pd.DataFrame(d).T, self.config["dump_path"] + "/test_report.xlsx", cls_name)


def summay_excel_report(df: pd.DataFrame, filename: os.PathLike, sheet: str, adjust: bool = True) -> None:
    if filename.exists():
        with pd.ExcelWriter(filename, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            df.to_excel(writer, sheet_name=sheet, index=False)
    else:
        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name=sheet, index=False)
    if adjust:
        adjust_column_width(filename)


def adjust_column_width(file_name: os.PathLike) -> None:
    wb = load_workbook(filename=file_name)
    for worksheets in wb.sheetnames:
        ws = wb[worksheets]
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column_letter].width = adjusted_width
    wb.save(filename=file_name)
