#!/usr/bin/env python
# -*-coding:utf-8 -*-
"""
@Project :   Study
@File    :   pandas_utils.py
@Time    :   2024-04-09 11:04
@Author  :   Wu Xiaomin <>
@Version :   1.0
@License :   (C)Copyright 2024, Wu Xiaomin
@Desc    :   pandas 相关工具类
"""
import os
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook


def summary_excel_report(df: pd.DataFrame, filename: Path, sheet: str, adjust: bool = True) -> None:
    """
    将 Dataframe 数据保存入 excel filename 的 sheet 中
    Args:
        df (): 要保存的数据
        filename ():  Excel 文件
        sheet ():  Excel Sheet name
        adjust ():  是否调整列宽, 如果为 True, 将根据内容调整列宽

    Returns:

    """
    if filename.exists():
        with pd.ExcelWriter(filename, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            df.to_excel(writer, sheet_name=sheet, index=False)
    else:
        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name=sheet, index=False)
    if adjust:
        adjust_column_width(filename)


def adjust_column_width(file_name: os.PathLike) -> None:
    """
    调整 Excel 文件的列宽
    Args:
        file_name ():  Excel 文件名称

    Returns:

    """
    wb = load_workbook(filename=file_name)
    for worksheets in wb.sheetnames:
        ws = wb[worksheets]
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column_letter].width = adjusted_width
    wb.save(filename=file_name)
