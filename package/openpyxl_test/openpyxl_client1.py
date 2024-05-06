#!/usr/bin/env python
# -*-coding:utf-8 -*-
"""
@Project :   Study
@File    :   openpyxl_client1.py
@Time    :   2024-05-03 20:36
@Author  :   Wu Xiaomin <>
@Version :   1.0
@License :   (C)Copyright 2024, Wu Xiaomin
@Desc    :   
"""
from openpyxl import Workbook
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles import Alignment
from openpyxl.drawing.image import Image
from openpyxl import load_workbook


def merge_cell():
    # 创建一个新的工作簿
    wb = Workbook()
    ws = wb.active
    # 合并A1到C3的单元格
    start_cell = 'A1'
    end_cell = 'F8'
    # 写入数据到合并的单元格
    ws[start_cell] = '合并后的数据'
    # 获取起始单元格的行索引和列索引
    start_row = int(start_cell[1:])
    start_col = column_index_from_string(start_cell[0])
    # 获取结束单元格的行索引和列索引
    end_row = int(end_cell[1:])
    end_col = column_index_from_string(end_cell[0])
    # 设置合并单元格的样式
    merge_range = f'{get_column_letter(start_col)}{start_row}:{get_column_letter(end_col)}{end_row}'
    ws.merge_cells(merge_range)
    # 针对每个合并的单元格应用对齐样式
    for row in ws[merge_range]:
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')
    # 保存工作簿到文件
    wb.save('merged_data.xlsx')


def add_image():
    work_book = Workbook()
    sheet = work_book.create_sheet(title="我的数据", index=0)  # title就是sheet名
    # 读入一张图片
    my_img = Image('./panda.png')
    # 将图片写入sheet中，A10是图片的插入位置
    # 调整表格列宽和行高
    sheet.column_dimensions["A"].width = 15
    sheet.row_dimensions[10].height = 90
    sheet.add_image(my_img, 'A10')
    work_book.save("我的数据.xlsx")


if __name__ == '__main__':
    # merge_cell()
    # add_image()
    wb = load_workbook(filename="我的数据.xlsx")
    ws = wb["我的数据"]
    print(ws)
    ws.merge_cells(start_row=1, end_row=7, start_column=2, end_column=2)
    print(ws.merged_cells.ranges)
    for m in ws.merged_cells.ranges:
        ws[str(m).split(":")[0]].alignment = Alignment(horizontal="center", vertical="center")
    wb.save("我的数据.xlsx")